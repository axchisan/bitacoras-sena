import re
import shutil
import zipfile
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XlImage
from app.config import get_settings

settings = get_settings()

TEMPLATE_PATH = Path("/app/templates/bitacora_base.xlsx")
OUTPUT_DIR = Path("/app/output")
_CLEAN_TEMPLATE: Path | None = None  # cached cleaned template

SHEET_NAME = "Formato Bitácora Individual"
ROW_BITACORA_NUMBER = 11
COL_BITACORA_NUMBER = 2
ROW_PERIOD = 11
COL_PERIOD = 5
ROW_DELIVERY_DATE = 71
COL_DELIVERY_DATE = 8

ACTIVITIES_START_ROW = 47
ACTIVITY_ROW_STEP = 2

COL_DESCRIPTION = 2
COL_COMPETENCIAS = 4
COL_START_DATE = 6
COL_END_DATE = 7
COL_EVIDENCE = 8
COL_OBSERVATIONS = 9


def _strip_worksheet_xml(txt: str) -> str:
    """Remove all elements from worksheet XML that cause openpyxl to hang on save."""
    # Remove drawing references
    txt = re.sub(r"<drawing[^/]*/?>", "", txt)
    txt = re.sub(r"<drawing[^>]*>.*?</drawing>", "", txt, flags=re.DOTALL)
    # Remove standard dataValidations block
    txt = re.sub(r"<dataValidations[^>]*>.*?</dataValidations>", "", txt, flags=re.DOTALL)
    # Remove conditional formatting (can be huge and slow)
    txt = re.sub(r"<conditionalFormatting[^>]*>.*?</conditionalFormatting>", "", txt, flags=re.DOTALL)
    # Remove extension lists (contains x14:dataValidations, sparklines, etc.)
    txt = re.sub(r"<extLst[^>]*>.*?</extLst>", "", txt, flags=re.DOTALL)
    # Remove table parts references
    txt = re.sub(r"<tableParts[^/]*/?>", "", txt)
    txt = re.sub(r"<tableParts[^>]*>.*?</tableParts>", "", txt, flags=re.DOTALL)
    # Remove OLE objects and controls
    txt = re.sub(r"<oleObjects[^>]*>.*?</oleObjects>", "", txt, flags=re.DOTALL)
    txt = re.sub(r"<controls[^>]*>.*?</controls>", "", txt, flags=re.DOTALL)
    # Remove legacy drawing references
    txt = re.sub(r"<legacyDrawing[^/]*/?>", "", txt)
    txt = re.sub(r"<legacyDrawingHF[^/]*/?>", "", txt)
    return txt


def _strip_drawings(src: Path, dst: Path) -> None:
    """
    Copy xlsx stripping DrawingML and other elements that cause openpyxl to
    hang on load or save. xlsx is a zip — we manipulate the XML directly.
    """
    with zipfile.ZipFile(src, "r") as zin, \
         zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:

        for item in zin.infolist():
            # Skip drawing binary files entirely
            if re.search(r"xl/(drawings|charts|chartsheets|diagrams)/", item.filename):
                continue
            # Skip VML drawings (legacy shapes)
            if item.filename.startswith("xl/drawings") or "vmlDrawing" in item.filename:
                continue
            # Skip media files (images embedded in drawings)
            if item.filename.startswith("xl/media/"):
                continue

            data = zin.read(item.filename)

            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                txt = data.decode("utf-8", errors="replace")
                txt = _strip_worksheet_xml(txt)
                data = txt.encode("utf-8")

            elif item.filename == "[Content_Types].xml":
                txt = data.decode("utf-8", errors="replace")
                # Remove drawing, chart, vml, and diagram content type overrides
                txt = re.sub(r'<Override[^>]*(drawing|chart|vml|diagram|Drawing|Chart)[^>]*/>', "", txt)
                data = txt.encode("utf-8")

            elif item.filename.endswith(".rels"):
                txt = data.decode("utf-8", errors="replace")
                # Remove relationships to drawings, charts, vml, media
                txt = re.sub(r'<Relationship[^>]*(drawing|chart|vml|media|diagram|Drawing|Chart)[^>]*/>', "", txt)
                data = txt.encode("utf-8")

            zout.writestr(item, data)


def _get_clean_template() -> Path:
    """Return a cached drawing-free copy of the template."""
    global _CLEAN_TEMPLATE
    if _CLEAN_TEMPLATE and _CLEAN_TEMPLATE.exists():
        return _CLEAN_TEMPLATE

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    clean = OUTPUT_DIR / "_bitacora_template_clean.xlsx"
    _strip_drawings(TEMPLATE_PATH, clean)
    _CLEAN_TEMPLATE = clean
    return clean


def _col_letter(col: int) -> str:
    """Convert 1-based column index to Excel letter (1→A, 8→H)."""
    result = ""
    while col:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _insert_evidence_images(ws, row: int, col: int, image_paths: list[str]) -> None:
    """Insert evidence images anchored to the evidence cell of each activity row."""
    col_letter = _col_letter(col)
    valid_paths = [p for p in image_paths if p and Path(p).exists()]
    if not valid_paths:
        return

    img_w, img_h = 160, 120  # pixels per image
    # Set row tall enough to show all stacked images (1 pt ≈ 1.33 px)
    ws.row_dimensions[row].height = max(90, len(valid_paths) * (img_h * 0.75 + 4))

    for idx, img_path in enumerate(valid_paths):
        try:
            xl_img = XlImage(img_path)
            xl_img.width = img_w
            xl_img.height = img_h
            # Simple string anchor: "H47", "H49", etc. openpyxl places the image
            # at the top-left corner of that cell. Multiple images stack.
            ws.add_image(xl_img, f"{col_letter}{row + idx}")
        except Exception:
            pass  # Never break the export if an image fails


def generate_excel(
    bitacora_number: int,
    period_start: date,
    period_end: date,
    activities: list[dict],
    delivery_date: date | None = None,
) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"Bitacora{bitacora_number}_DuvanYairArciniegas.xlsx"

    clean_template = _get_clean_template()
    shutil.copy2(clean_template, output_path)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb[SHEET_NAME]

    ws.cell(row=ROW_BITACORA_NUMBER, column=COL_BITACORA_NUMBER).value = str(bitacora_number)

    period_label = (
        f"Desde {period_start.strftime('%d/%m/%Y')} "
        f"hasta {period_end.strftime('%d/%m/%Y')}"
    )
    ws.cell(row=ROW_PERIOD, column=COL_PERIOD).value = period_label

    if delivery_date:
        ws.cell(row=ROW_DELIVERY_DATE, column=COL_DELIVERY_DATE).value = (
            delivery_date.strftime("%d/%m/%Y")
        )

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for i, activity in enumerate(activities):
        row = ACTIVITIES_START_ROW + (i * ACTIVITY_ROW_STEP)

        desc_cell = ws.cell(row=row, column=COL_DESCRIPTION)
        desc_cell.value = f"{activity.get('title', '')}\n{activity.get('description', '')}"
        desc_cell.alignment = wrap_alignment

        comp_cell = ws.cell(row=row, column=COL_COMPETENCIAS)
        comp_cell.value = activity.get("competencias", "")
        comp_cell.alignment = wrap_alignment

        if activity.get("start_date"):
            start = activity["start_date"]
            if isinstance(start, str):
                from datetime import datetime as dt
                start = dt.strptime(start, "%Y-%m-%d").date()
            ws.cell(row=row, column=COL_START_DATE).value = start.strftime("%d/%m/%y")

        if activity.get("end_date"):
            end = activity["end_date"]
            if isinstance(end, str):
                from datetime import datetime as dt
                end = dt.strptime(end, "%Y-%m-%d").date()
            ws.cell(row=row, column=COL_END_DATE).value = end.strftime("%d/%m/%y")

        ev_cell = ws.cell(row=row, column=COL_EVIDENCE)
        ev_cell.value = activity.get("evidence_description", "")
        ev_cell.alignment = wrap_alignment

        obs_cell = ws.cell(row=row, column=COL_OBSERVATIONS)
        obs_cell.value = activity.get("observations", "")
        obs_cell.alignment = wrap_alignment

        # Insertar imágenes de evidencia en la celda correspondiente
        evidence_images = activity.get("evidence_images", [])
        if evidence_images:
            _insert_evidence_images(ws, row, COL_EVIDENCE, evidence_images)

    wb.save(str(output_path))
    wb.close()

    return output_path
